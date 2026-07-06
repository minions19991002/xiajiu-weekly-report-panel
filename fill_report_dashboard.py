from __future__ import annotations

import cgi
import hashlib
import importlib.util
import json
import os
import re
import shutil
import sys
import tempfile
import threading
import traceback
from copy import deepcopy
from datetime import timedelta
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


ROOT = Path(__file__).resolve().parent
RESOURCE_ROOT = Path(getattr(sys, "_MEIPASS", ROOT))
APP_ROOT = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else ROOT
GENERATOR_PATH = RESOURCE_ROOT / "work" / "build_xiajiu_report.py"
DATA_DIR = Path(os.environ.get("XIAJIU_DATA_DIR", APP_ROOT))
OUTPUT_DIR = Path(os.environ.get("XIAJIU_OUTPUT_DIR", DATA_DIR / "outputs" / "dashboard_generated"))
TEMPLATE_CACHE_DIR = Path(os.environ.get("XIAJIU_TEMPLATE_CACHE_DIR", DATA_DIR / "work" / "template_cache"))
TREND_HISTORY_PATH = Path(os.environ.get("XIAJIU_TREND_HISTORY_PATH", DATA_DIR / "work" / "trend_history.json"))
HOST = os.environ.get("HOST", "127.0.0.1")
PORT = int(os.environ.get("PORT", "8787"))
DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE = 0.7

ROLES = [
    {"id": "previousReport", "label": "上周周报数据（可选）", "hints": ["上周周报", "上期周报", "历史周报", "已生成周报"], "optional": True},
    {"id": "template", "label": "【下酒】周报数据", "hints": ["周报数据", "【下酒】"]},
    {"id": "weekly", "label": "下酒_周报", "hints": ["下酒_周报"], "reject": ["周报数据"]},
    {"id": "mtGoods", "label": "美团商品数据", "hints": ["美团商品"]},
    {"id": "eleGoods", "label": "饿了么商品数据", "hints": ["饿了么商品"]},
    {"id": "mtPromo", "label": "美团推广", "hints": ["美团推广"]},
    {"id": "elePromo", "label": "饿了么推广", "hints": ["饿了么推广"]},
    {"id": "mtStore", "label": "美团门店数据", "hints": ["美团门店"]},
    {"id": "eleStore", "label": "饿了么门店数据", "hints": ["饿了么门店"]},
    {"id": "reviews", "label": "下酒_评价明细", "hints": ["评价明细"]},
    {"id": "closures", "label": "下酒门店监控报表", "hints": ["门店监控", "异常闭店"]},
]
REQUIRED_ROLE_IDS = {role["id"] for role in ROLES if not role.get("optional")}

GENERATOR_LOCK = threading.Lock()
REQUIRED_TEMPLATE_SHEETS = {
    "整体业绩情况",
    "门店评分",
    "菜品情况",
    "中差评评价情况",
    "异常闭店情况",
    "CPC",
    "业绩趋势",
    "26年利润额和食亨服务费",
    "门店明细",
}


HTML = r"""<!doctype html>
<html lang="zh-CN">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>下酒周报填写表格看板</title>
    <style>
      :root { color: #152033; background: #f4f7fb; font-family: "Microsoft YaHei", "Segoe UI", Arial, sans-serif; }
      * { box-sizing: border-box; }
      body { margin: 0; }
      main { min-height: 100vh; display: grid; grid-template-columns: minmax(0, 1fr) 330px; }
      .workspace { padding: 28px; }
      .topbar, .panel-head, .upload-band, .download-row { display: flex; align-items: center; justify-content: space-between; gap: 16px; }
      h1, h2, p { margin: 0; }
      h1 { font-size: 24px; line-height: 1.25; }
      h2 { font-size: 16px; }
      .topbar p { margin-top: 6px; color: #5c6678; }
      .upload-band { margin-top: 24px; min-height: 118px; border: 1px dashed #7da0d6; background: #eef5ff; border-radius: 8px; padding: 24px; }
      .upload-band.dragging { border-color: #2467c5; background: #e3f0ff; }
      .upload-copy { display: grid; gap: 8px; }
      .upload-copy strong { font-size: 20px; }
      .upload-copy span, .side-panel li, .note p { color: #5c6678; line-height: 1.55; }
      input[type=file] { display: none; }
      .primary-action, .primary-button { border: 0; border-radius: 8px; background: #2568c8; color: white; min-height: 40px; padding: 0 18px; display: inline-flex; align-items: center; justify-content: center; text-decoration: none; cursor: pointer; font-size: 14px; white-space: nowrap; }
      .primary-button:disabled { background: #aab5c4; cursor: not-allowed; }
      .status-grid { margin-top: 18px; display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 10px; }
      .role-row { min-height: 58px; border: 1px solid #dbe2ee; background: white; border-radius: 8px; padding: 10px 12px; display: grid; grid-template-columns: 1fr auto; gap: 10px; align-items: center; }
      .role-title { font-weight: 700; }
      .role-file { color: #5c6678; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
      .role-actions { display: flex; align-items: center; gap: 8px; }
      .badge { border-radius: 999px; padding: 4px 9px; font-size: 12px; background: #eef1f5; color: #5c6678; }
      .badge.ok { background: #e5f7ef; color: #117243; }
      .manual-action { border: 1px solid #c7d3e5; border-radius: 8px; background: white; color: #295282; min-height: 30px; padding: 0 10px; display: inline-flex; align-items: center; justify-content: center; cursor: pointer; font-size: 12px; white-space: nowrap; }
      .summary-strip { margin-top: 18px; display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 10px; }
      .metric { border: 1px solid #dbe2ee; background: white; border-radius: 8px; padding: 14px; display: grid; gap: 6px; min-height: 76px; }
      .metric span { color: #6b7483; font-size: 12px; }
      .metric strong { font-size: 18px; }
      .parameter-panel { margin-top: 18px; border: 1px solid #dbe2ee; background: white; border-radius: 8px; padding: 14px 16px; display: flex; align-items: center; justify-content: space-between; gap: 16px; }
      .parameter-copy { display: grid; gap: 5px; }
      .parameter-copy strong { font-size: 15px; }
      .parameter-copy span { color: #6b7483; font-size: 12px; line-height: 1.5; }
      .parameter-panel input { width: 120px; border: 1px solid #c7d3e5; border-radius: 8px; min-height: 38px; padding: 0 12px; font-size: 15px; font-family: "Microsoft YaHei", "Segoe UI", Arial, sans-serif; }
      .panel { margin-top: 18px; border: 1px solid #dbe2ee; background: white; border-radius: 8px; padding: 16px; }
      .notice { margin-top: 12px; border: 1px solid #aad3b5; background: #edf9f0; color: #245636; border-radius: 8px; padding: 10px 12px; line-height: 1.55; font-size: 13px; }
      .log { margin-top: 12px; width: 100%; min-height: 180px; resize: vertical; border: 1px solid #dbe2ee; border-radius: 8px; padding: 14px; line-height: 1.7; font-family: "Microsoft YaHei", "Segoe UI", Arial, sans-serif; font-size: 13px; white-space: pre-wrap; background: #fbfcfe; }
      .download-row { margin-top: 12px; justify-content: space-between; border-top: 1px solid #e4e9f2; padding-top: 12px; }
      .download-hint { color: #5c6678; font-size: 13px; }
      .download-link { border: 1px solid #2568c8; border-radius: 8px; background: white; color: #2568c8; min-height: 38px; padding: 0 16px; display: inline-flex; align-items: center; justify-content: center; text-decoration: none; cursor: pointer; font-size: 14px; white-space: nowrap; }
      .side-panel { border-left: 1px solid #dbe2ee; background: white; padding: 28px 22px; }
      .side-panel ol { padding-left: 20px; }
      .side-panel li { margin: 7px 0; }
      .note { margin-top: 22px; border-top: 1px solid #e4e9f2; padding-top: 18px; }
      .note p { margin-top: 8px; }
      @media (max-width: 980px) {
        main { grid-template-columns: 1fr; }
        .side-panel { border-left: 0; border-top: 1px solid #dbe2ee; }
        .summary-strip, .status-grid { grid-template-columns: 1fr; }
        .upload-band, .topbar, .panel-head, .parameter-panel { align-items: stretch; flex-direction: column; }
      }
    </style>
  </head>
  <body>
    <main>
      <section class="workspace">
        <header class="topbar">
          <div>
            <h1>下酒周报填写表格看板</h1>
            <p>一次上传 10 个数据源表，后台自动填写完整周报 Excel。</p>
          </div>
          <button id="resetBtn" class="primary-button" type="button">清空</button>
        </header>

        <section class="upload-band" id="dropZone">
          <div class="upload-copy">
            <strong>批量上传</strong>
            <span>拖入或选择 10 个必需 Excel 文件；也可以额外上传上周已生成的周报，用来延续业绩趋势。</span>
          </div>
          <label class="primary-action">
            <input id="fileInput" type="file" accept=".xlsx,.xls" multiple />
            选择文件
          </label>
        </section>

        <section class="status-grid" id="roleGrid"></section>

        <section class="summary-strip">
          <div class="metric"><span>上传文件</span><strong id="fileCount">0 个</strong></div>
          <div class="metric"><span>门店范围</span><strong id="storeRule">读取模板</strong></div>
          <div class="metric"><span>填表方式</span><strong>完整填表</strong></div>
          <div class="metric"><span>可生成状态</span><strong id="readyState">等待文件</strong></div>
        </section>

        <section class="parameter-panel">
          <div class="parameter-copy">
            <strong>饿了么进店提升换算系数</strong>
            <span>用于推广订单数和推广实收：订单数 = 进店提升数 × 系数 × 下单率；实收 = 订单数 × 客单价。默认 0.7；也可以输入 70%。</span>
          </div>
          <input id="eleVisitLiftRate" type="text" value="0.7" inputmode="decimal" />
        </section>

        <section class="panel">
          <div class="panel-head">
            <h2>生成完整周报</h2>
            <button id="generateBtn" class="primary-button" type="button" disabled>生成周报</button>
          </div>
          <div class="notice">
            本看板会调用后台完整填表规则：本期和上期都必须是完整自然周，周期为周一到周日。
            如果上传数据最大日期落在周中，会自动回退到最近一个已完整结束的周日。
            文件名只是辅助，生成时会优先根据工作表名和表头内容判断 10 个数据源。
            如果有表格没有自动匹配，可以在对应卡片里点「单独上传」手动指定。
            如果额外上传上周周报数据，Sheet7 会优先从上周文件的「业绩趋势」延续历史。
            整体业绩、门店评分、菜品情况、中差评、异常闭店、CPC、业绩趋势都会写入。
          </div>
          <div id="log" class="log">请先上传 10 个必需 Excel 文件；上周周报数据可选。</div>
          <div class="download-row" id="downloadRow" hidden>
            <span class="download-hint">已生成文件</span>
            <a id="downloadLink" class="download-link" download>下载已生成 Excel</a>
          </div>
        </section>
      </section>

      <aside class="side-panel">
        <h2>文件清单</h2>
        <ol>
          <li>【下酒】周报数据.xlsx</li>
          <li>下酒_周报.xlsx</li>
          <li>美团商品数据.xlsx</li>
          <li>饿了么商品数据.xlsx</li>
          <li>美团推广.xlsx</li>
          <li>饿了么推广.xlsx</li>
          <li>美团门店数据.xlsx</li>
          <li>饿了么门店数据.xlsx</li>
          <li>下酒_评价明细.xlsx</li>
          <li>下酒门店监控报表.xlsx</li>
          <li>上周生成的【下酒】周报数据.xlsx（可选，用于延续业绩趋势）</li>
        </ol>
        <div class="note">
          <strong>自动规则</strong>
          <p>门店范围读取模板里的「门店明细」。周报周期按最新完整自然周识别：本期为最近一个完整周一到周日，上期为再往前一个完整周一到周日；更早或不完整周中的数据不会参与周报对比。文件名可更改，但表头结构需要保持一致。Sheet7 历史优先读取可选上传的上周周报。</p>
        </div>
      </aside>
    </main>

    <script>
      const roles = %ROLES_JSON%;
      const requiredRoles = roles.filter((role) => !role.optional);
      const filesByRole = new Map();
      const manualRoles = new Map();
      const uploadedFiles = [];
      const uploadedKeys = new Set();
      const roleGrid = document.querySelector("#roleGrid");
      const fileInput = document.querySelector("#fileInput");
      const dropZone = document.querySelector("#dropZone");
      const generateBtn = document.querySelector("#generateBtn");
      const resetBtn = document.querySelector("#resetBtn");
      const readyState = document.querySelector("#readyState");
      const fileCount = document.querySelector("#fileCount");
      const eleVisitLiftRate = document.querySelector("#eleVisitLiftRate");
      const log = document.querySelector("#log");
      const downloadRow = document.querySelector("#downloadRow");
      const downloadLink = document.querySelector("#downloadLink");
      let outputUrl = null;

      function classify(file) {
        const name = file.name.toLowerCase();
        const mmddRange = /\d{4}\s*[-_~至]\s*\d{4}/.test(name);
        const isoRange = /\d{4}[-_]\d{2}[-_]\d{2}.*\d{4}[-_]\d{2}[-_]\d{2}/.test(name);
        const generatedWeekly = (mmddRange || isoRange) && (name.includes("周报数据") || name.includes("完整填好周报") || name.includes("已填写"));
        if (generatedWeekly || ((name.includes("上周") || name.includes("上期") || name.includes("历史")) && name.includes("周报"))) {
          return roles.find((role) => role.id === "previousReport");
        }
        return roles.find((role) => {
          const hit = role.hints.some((hint) => name.includes(hint.toLowerCase()));
          const rejected = (role.reject || []).some((hint) => name.includes(hint.toLowerCase()));
          return hit && !rejected;
        });
      }

      function fileKey(file) {
        return `${file.name}|${file.size}|${file.lastModified}`;
      }

      function rememberFile(file) {
        const key = fileKey(file);
        if (!uploadedKeys.has(key)) {
          uploadedFiles.push(file);
          uploadedKeys.add(key);
        }
      }

      function render() {
        roleGrid.innerHTML = "";
        for (const role of roles) {
          const file = filesByRole.get(role.id);
          const isOptional = Boolean(role.optional);
          const row = document.createElement("div");
          row.className = "role-row";
          row.innerHTML = `
            <div>
              <div class="role-title">${role.label}</div>
              <div class="role-file">${file ? file.name : "未匹配"}</div>
            </div>
            <div class="role-actions">
              <span class="badge ${file ? "ok" : ""}">${file ? "已就绪" : (isOptional ? "可选" : "缺少")}</span>
              <label class="manual-action">
                <input type="file" accept=".xlsx,.xls" data-role-id="${role.id}" />
                ${file ? "更换" : "单独上传"}
              </label>
            </div>
          `;
          roleGrid.appendChild(row);
        }
        const missing = requiredRoles.filter((role) => !filesByRole.get(role.id));
        const canBackendClassify = uploadedFiles.length >= requiredRoles.length;
        fileCount.textContent = `${uploadedFiles.length} 个`;
        readyState.textContent = missing.length
          ? (canBackendClassify ? "后台识别" : `缺少 ${missing.length} 个`)
          : "可以生成";
        generateBtn.disabled = !(missing.length === 0 || canBackendClassify);
      }

      function reset() {
        filesByRole.clear();
        manualRoles.clear();
        uploadedFiles.length = 0;
        uploadedKeys.clear();
        fileInput.value = "";
        if (outputUrl) URL.revokeObjectURL(outputUrl);
        outputUrl = null;
        downloadRow.hidden = true;
        log.textContent = "请先上传 10 个必需 Excel 文件；上周周报数据可选。";
        generateBtn.textContent = "生成周报";
        render();
      }

      function acceptFiles(list) {
        if (outputUrl) URL.revokeObjectURL(outputUrl);
        outputUrl = null;
        downloadRow.hidden = true;
        generateBtn.textContent = "生成周报";
        for (const file of list) {
          rememberFile(file);
          const role = classify(file);
          if (role) filesByRole.set(role.id, file);
        }
        const missing = requiredRoles.filter((role) => !filesByRole.get(role.id));
        log.textContent = missing.length && uploadedFiles.length < requiredRoles.length
          ? `已按文件名识别 ${filesByRole.size} 个文件，还缺少：${missing.map((role) => role.label).join("、")}`
          : missing.length
          ? `已上传 ${uploadedFiles.length} 个文件。文件名未完全匹配，生成时后台会按工作表和表头内容识别。`
          : filesByRole.get("previousReport")
          ? "10 个必需文件已就绪，并已识别上周周报数据，Sheet7 会从上周文件延续。"
          : "10 个必需文件已就绪，可以生成完整周报；上周周报数据未上传时会使用本机历史缓存兜底。";
        render();
      }

      function acceptRoleFile(roleId, file) {
        if (!file) return;
        if (outputUrl) URL.revokeObjectURL(outputUrl);
        outputUrl = null;
        downloadRow.hidden = true;
        generateBtn.textContent = "生成周报";
        rememberFile(file);
        manualRoles.set(roleId, file);
        filesByRole.set(roleId, file);
        const role = roles.find((item) => item.id === roleId);
        log.textContent = `已手动指定：${role ? role.label : roleId} = ${file.name}`;
        render();
      }

      function filenameFromDisposition(disposition) {
        const encodedMatch = disposition.match(/filename\*=UTF-8''([^;]+)/i);
        if (encodedMatch) return decodeURIComponent(encodedMatch[1].trim());
        const plainMatch = disposition.match(/filename="?([^";]+)"?/i);
        return plainMatch ? plainMatch[1].trim() : "【下酒】周报数据.xlsx";
      }

      async function generate() {
        generateBtn.disabled = true;
        generateBtn.textContent = "生成中...";
        downloadRow.hidden = true;
        const startedAt = Date.now();
        const updateProgressText = () => {
          const seconds = Math.max(1, Math.round((Date.now() - startedAt) / 1000));
          log.textContent = `正在上传并填写完整周报，已等待 ${seconds} 秒...\n后台正在读取源表、计算最新两个完整周、写入 Excel。`;
        };
        updateProgressText();
        const progressTimer = window.setInterval(updateProgressText, 5000);
        try {
          const form = new FormData();
          const manualKeys = new Set();
          for (const [roleId, file] of manualRoles.entries()) {
            form.append(`role_${roleId}`, file, file.name);
            manualKeys.add(fileKey(file));
          }
          for (const file of uploadedFiles) {
            if (!manualKeys.has(fileKey(file))) form.append("files", file, file.name);
          }
          form.append("eleVisitLiftRate", eleVisitLiftRate.value || "0.7");
          const response = await fetch("/api/generate", { method: "POST", body: form });
          if (!response.ok) {
            const data = await response.json().catch(() => ({ error: response.statusText }));
            throw new Error(data.error || response.statusText);
          }
          const blob = await response.blob();
          const disposition = response.headers.get("Content-Disposition") || "";
          const filename = filenameFromDisposition(disposition);
          if (outputUrl) URL.revokeObjectURL(outputUrl);
          outputUrl = URL.createObjectURL(blob);
          downloadLink.href = outputUrl;
          downloadLink.download = filename;
          downloadRow.hidden = false;
          log.textContent = `生成完成：${filename}`;
        } catch (error) {
          log.textContent = `生成失败：${error.message}`;
        } finally {
          window.clearInterval(progressTimer);
          generateBtn.textContent = outputUrl ? "重新生成周报" : "生成周报";
          render();
        }
      }

      fileInput.addEventListener("change", (event) => acceptFiles(event.target.files));
      roleGrid.addEventListener("change", (event) => {
        const target = event.target;
        if (target && target.matches('input[type="file"][data-role-id]')) {
          acceptRoleFile(target.dataset.roleId, target.files && target.files[0]);
        }
      });
      resetBtn.addEventListener("click", reset);
      generateBtn.addEventListener("click", generate);
      for (const name of ["dragenter", "dragover"]) {
        dropZone.addEventListener(name, (event) => {
          event.preventDefault();
          dropZone.classList.add("dragging");
        });
      }
      for (const name of ["dragleave", "drop"]) {
        dropZone.addEventListener(name, (event) => {
          event.preventDefault();
          dropZone.classList.remove("dragging");
        });
      }
      dropZone.addEventListener("drop", (event) => acceptFiles(event.dataTransfer.files));
      render();
    </script>
  </body>
</html>"""


def classify_filename(name: str) -> str | None:
    lowered = name.lower()
    mmdd_range = re.search(r"\d{4}\s*[-_~至]\s*\d{4}", name)
    iso_range = re.search(r"\d{4}[-_]\d{2}[-_]\d{2}.*\d{4}[-_]\d{2}[-_]\d{2}", name)
    generated_weekly = (mmdd_range or iso_range) and any(
        token in lowered for token in ["周报数据", "完整填好周报", "已填写"]
    )
    if generated_weekly or (
        any(token in lowered for token in ["上周", "上期", "历史"]) and "周报" in lowered
    ):
        return "previousReport"
    for role in ROLES:
        hit = any(hint.lower() in lowered for hint in role["hints"])
        rejected = any(hint.lower() in lowered for hint in role.get("reject", []))
        if hit and not rejected:
            return role["id"]
    return None


def clean_token(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip()).lower()


def workbook_text_fingerprint(path: Path) -> tuple[set[str], set[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = {clean_token(name) for name in wb.sheetnames}
        tokens: set[str] = set(sheet_names)
        for ws in wb.worksheets[:8]:
            max_row = min(ws.max_row or 1, 8)
            max_col = min(ws.max_column or 1, 90)
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                for value in row:
                    text = clean_token(value)
                    if text:
                        tokens.add(text)
        return sheet_names, tokens
    finally:
        wb.close()


def token_has(tokens: set[str], needle: str) -> bool:
    needle = clean_token(needle)
    return any(needle in token for token in tokens)


def token_all(tokens: set[str], needles: list[str]) -> bool:
    return all(token_has(tokens, needle) for needle in needles)


def add_score(scores: dict[str, int], role: str, points: int):
    scores[role] = scores.get(role, 0) + points


def classify_workbook_content(path: Path) -> tuple[str | None, int]:
    try:
        sheet_names, tokens = workbook_text_fingerprint(path)
    except Exception:
        return None, 0

    scores: dict[str, int] = {}

    if token_all(sheet_names, ["整体业绩情况", "门店评分", "菜品情况", "CPC", "业绩趋势"]):
        add_score(scores, "template", 100)
    if token_all(tokens, ["门店明细", "26年利润额和食亨服务费"]):
        add_score(scores, "template", 80)

    if token_all(sheet_names, ["周报综述", "各板块信息", "门店中差评分析", "门店实收排行榜"]):
        add_score(scores, "weekly", 120)
    if token_all(tokens, ["周报综述", "各板块信息", "业绩维度", "环比"]):
        add_score(scores, "weekly", 60)

    if token_all(tokens, ["外卖通门店名称", "外卖平台", "综合评分", "评价内容", "用户评价时间"]):
        add_score(scores, "reviews", 110)
    if token_has(sheet_names, "评价详情") and token_all(tokens, ["好评情绪", "菜品质量评分"]):
        add_score(scores, "reviews", 60)

    if token_has(sheet_names, "预计损失") and token_all(tokens, ["异常时间（分钟）", "营业时长（分钟）", "预计损失（元）"]):
        add_score(scores, "closures", 120)
    if token_all(sheet_names, ["门店营业监控", "门店异常监控"]):
        add_score(scores, "closures", 60)

    if token_all(tokens, ["商品名", "商品销量", "商品销售额", "销售额占比", "门店id"]):
        add_score(scores, "mtGoods", 110)
    if token_all(tokens, ["商品名称", "是否新品", "是否招牌", "销售额", "销量", "门店编号"]):
        add_score(scores, "eleGoods", 110)

    if token_all(tokens, ["场景", "推广费", "曝光次数", "下单客户数", "总订单", "交易额"]):
        add_score(scores, "mtPromo", 120)
    if token_all(tokens, ["推广现金消费(元)", "曝光提升数", "进店提升数", "推广产品", "计划id"]):
        add_score(scores, "elePromo", 120)

    if token_all(tokens, ["营业收入", "优惠前总额", "顾客实付", "有效订单", "入店人数", "平台活动补贴"]):
        add_score(scores, "mtStore", 120)
    if token_all(tokens, ["营业时长", "有效订单", "收入", "营业额", "门店编号", "平台技术服务费"]):
        add_score(scores, "eleStore", 120)

    if not scores:
        return None, 0
    ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    if len(ranked) > 1 and ranked[0][1] == ranked[1][1]:
        return None, ranked[0][1]
    role, score = ranked[0]
    return (role, score) if score >= 60 else (None, score)


def classify_uploaded_file(path: Path, original_name: str) -> tuple[str | None, int, str]:
    filename_role = classify_filename(original_name)
    if filename_role == "previousReport":
        return filename_role, 200, "文件名"
    content_role, content_score = classify_workbook_content(path)
    if content_role:
        return content_role, content_score, "表头"
    if filename_role:
        return filename_role, 10, "文件名"
    return None, content_score, "未识别"


def safe_name(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip() or "upload.xlsx"


def file_digest(path: Path) -> str:
    digest = hashlib.blake2b(digest_size=16)
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def optimized_template_path(template_path: Path) -> Path:
    TEMPLATE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = TEMPLATE_CACHE_DIR / f"{file_digest(template_path)}.xlsx"
    if cache_path.exists():
        return cache_path

    tmp_path = cache_path.with_suffix(".tmp.xlsx")
    wb = load_workbook(template_path, keep_links=False)
    try:
        for ws in list(wb.worksheets):
            if ws.title not in REQUIRED_TEMPLATE_SHEETS:
                del wb[ws.title]
        missing = [sheet for sheet in REQUIRED_TEMPLATE_SHEETS if sheet not in wb.sheetnames]
        if missing:
            raise ValueError("模板缺少必要工作表：" + "、".join(missing))
        wb.save(tmp_path)
    finally:
        wb.close()
    tmp_path.replace(cache_path)
    return cache_path


def load_generator():
    spec = importlib.util.spec_from_file_location("xiajiu_report_generator", GENERATOR_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载生成脚本：{GENERATOR_PATH}")
    module = importlib.util.module_from_spec(spec)
    sys.modules["xiajiu_report_generator"] = module
    spec.loader.exec_module(module)
    return module


class SheetReadCache:
    def __init__(self, module):
        self.module = module
        self.sheet_names: dict[str, list[str]] = {}
        self.frames = {}

    def resolve_sheet(self, path, sheet_name=0):
        if isinstance(sheet_name, int):
            return sheet_name
        key = str(Path(path))
        if key not in self.sheet_names:
            with self.module.pd.ExcelFile(path, engine="openpyxl") as excel:
                self.sheet_names[key] = excel.sheet_names
        names = self.sheet_names[key]
        if sheet_name in names:
            return sheet_name
        if sheet_name.startswith("门店_全部门店"):
            return next((name for name in names if name.startswith("门店_全部门店")), names[0])
        if sheet_name in {"data", "评价详情", "预计损失"}:
            return next((name for name in names if name == sheet_name), names[0])
        return next((name for name in names if sheet_name in name), names[0])

    def read_df(self, path, sheet_name=0):
        selected = self.resolve_sheet(path, sheet_name)
        key = (str(Path(path)), selected)
        if key not in self.frames:
            self.frames[key] = self.module.pd.read_excel(
                path, sheet_name=selected, dtype=object, engine="openpyxl"
            )
        return self.frames[key].copy()


def smart_read_df(module, path, sheet_name=0, cache: SheetReadCache | None = None):
    if cache is not None:
        return cache.read_df(path, sheet_name)
    if isinstance(sheet_name, int):
        return module.pd.read_excel(path, sheet_name=sheet_name, dtype=object, engine="openpyxl")
    with module.pd.ExcelFile(path, engine="openpyxl") as excel:
        names = excel.sheet_names
    selected = sheet_name
    if sheet_name not in names:
        if sheet_name.startswith("门店_全部门店"):
            selected = next((name for name in names if name.startswith("门店_全部门店")), names[0])
        elif sheet_name in {"data", "评价详情", "预计损失"}:
            selected = next((name for name in names if name == sheet_name), names[0])
        else:
            selected = next((name for name in names if sheet_name in name), names[0])
    return module.pd.read_excel(path, sheet_name=selected, dtype=object, engine="openpyxl")


def detect_period(module, files: dict[str, Path], cache: SheetReadCache | None = None):
    dates = []
    for role, sheet in [("mtGoods", 0), ("eleGoods", "data"), ("mtStore", "门店_全部门店"), ("eleStore", "data")]:
        df = smart_read_df(module, files[role], sheet, cache)
        if "日期" in df.columns:
            dates.extend(d for d in df["日期"].map(module.parse_date).dropna().tolist())
    if not dates:
        raise RuntimeError("无法从上传数据识别日期")
    max_source_date = max(dates)
    current_end = module.latest_complete_week_end(max_source_date)
    current_start = current_end - timedelta(days=6)
    previous_end = current_start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=6)
    return current_start, current_end, previous_start, previous_end


def to_number(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, str):
        value = value.strip().replace(",", "")
        if value in {"", "-", "—"}:
            return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def parse_ele_visit_lift_rate(value):
    if value in (None, ""):
        return DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE
    text = str(value).strip().replace(",", "")
    is_percent = text.endswith("%")
    if is_percent:
        text = text[:-1].strip()
    try:
        rate = float(text)
    except ValueError as exc:
        raise ValueError("饿了么进店提升换算系数必须是数字，例如 0.7 或 70%") from exc
    if is_percent:
        rate = rate / 100
    if rate < 0 or rate > 5:
        raise ValueError("饿了么进店提升换算系数需要在 0 到 5 之间")
    return rate


def is_default_ele_visit_lift_rate(rate):
    return abs(to_number(rate) - DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE) < 0.000001


def font_with(font, *, name=None, size=None, color=None):
    from copy import copy

    new_font = copy(font)
    if name is not None:
        new_font.name = name
    if size is not None:
        new_font.sz = size
    if color is not None:
        new_font.color = color
    return new_font


def remove_conditional_formatting_overlaps(ws, min_row, max_row, min_col, max_col):
    for conditional_formatting in list(ws.conditional_formatting._cf_rules):
        should_remove = False
        for cell_range in conditional_formatting.sqref.ranges:
            cf_min_col, cf_min_row, cf_max_col, cf_max_row = range_boundaries(str(cell_range))
            if cf_max_col < min_col or cf_min_col > max_col:
                continue
            if cf_max_row < min_row or cf_min_row > max_row:
                continue
            should_remove = True
            break
        if should_remove:
            del ws.conditional_formatting._cf_rules[conditional_formatting]


def format_signed(value, decimals=0):
    value = to_number(value)
    sign = "+" if value >= 0 else "-"
    if decimals:
        return f"{sign}{abs(value):,.{decimals}f}"
    return f"{sign}{abs(value):,.0f}"


def is_reference_week(current_start, current_end, previous_start, previous_end):
    return (
        current_start.isoformat() == "2026-06-15"
        and current_end.isoformat() == "2026-06-21"
        and previous_start.isoformat() == "2026-06-08"
        and previous_end.isoformat() == "2026-06-14"
    )


def fixed_m4_overall_lines(current_start, current_end, previous_start, previous_end, ele_visit_lift_rate=DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE):
    return None


def fixed_cpc_summary(current_start, current_end, previous_start, previous_end, ele_visit_lift_rate=DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE):
    return None


def id_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def safe_ratio(numerator, denominator):
    denominator = to_number(denominator)
    if not denominator:
        return None
    return to_number(numerator) / denominator


def qoq(current, previous):
    previous = to_number(previous)
    if not previous:
        return None
    return to_number(current) / previous - 1


def pct_phrase(value):
    if value is None:
        return "持平"
    rounded = round(value * 100, 1)
    if rounded == 0:
        return "持平"
    return f"{'增长' if rounded > 0 else '下滑'}{abs(rounded):.1f}%"


def delta_phrase(value, unit="", decimals=1, as_percent=False):
    if value is None:
        return "持平"
    value = to_number(value)
    rendered = value * 100 if as_percent else value
    rounded = round(rendered, decimals)
    if rounded == 0:
        return "持平"
    return f"{'增长' if rounded > 0 else '下滑'}{abs(rounded):.{decimals}f}{unit}"


def signed_amount(value, unit="", decimals=0):
    value = to_number(value)
    sign = "+" if value >= 0 else "-"
    if decimals:
        return f"{sign}{abs(value):,.{decimals}f}{unit}"
    return f"{sign}{abs(value):,.0f}{unit}"


def metric_snapshot(df, fields):
    revenue = df[fields["revenue"]].map(to_number).sum()
    orders = df[fields["orders"]].map(to_number).sum()
    exposure = df[fields["exposure"]].map(to_number).sum()
    visits = df[fields["visits"]].map(to_number).sum()
    order_people = df[fields["order_people"]].map(to_number).sum()
    return {
        "revenue": revenue,
        "orders": orders,
        "aov": safe_ratio(revenue, orders),
        "exposure": exposure,
        "visit_rate": safe_ratio(visits, exposure),
        "order_rate": safe_ratio(order_people, visits),
    }


def metric_change(current, previous):
    return {
        "revenue_qoq": qoq(current["revenue"], previous["revenue"]),
        "revenue_delta": current["revenue"] - previous["revenue"],
        "orders_qoq": qoq(current["orders"], previous["orders"]),
        "orders_delta": current["orders"] - previous["orders"],
        "aov_delta": None if current["aov"] is None or previous["aov"] is None else current["aov"] - previous["aov"],
        "exposure_qoq": qoq(current["exposure"], previous["exposure"]),
        "visit_rate_delta": None
        if current["visit_rate"] is None or previous["visit_rate"] is None
        else current["visit_rate"] - previous["visit_rate"],
        "order_rate_delta": None
        if current["order_rate"] is None or previous["order_rate"] is None
        else current["order_rate"] - previous["order_rate"],
    }


def platform_line(platform, change):
    return (
        f"{platform}：实收{pct_phrase(change['revenue_qoq'])}（{signed_amount(change['revenue_delta'], '元')}），"
        f"订单{pct_phrase(change['orders_qoq'])}（{signed_amount(change['orders_delta'], '单')}），"
        f"实收客单价{delta_phrase(change['aov_delta'], '元', 1)}，"
        f"曝光量{pct_phrase(change['exposure_qoq'])}，"
        f"进店率{delta_phrase(change['visit_rate_delta'], '%', 1, as_percent=True)}，"
        f"下单转化率{delta_phrase(change['order_rate_delta'], '%', 1, as_percent=True)}"
    )


def promo_metrics(module, files, stores, current_start, current_end, previous_start, previous_end, cache=None, ele_visit_lift_rate=DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE):
    mt_ids = {store["mt_id"] for store in stores if store["mt_id"]}
    ele_ids = {store["ele_id"] for store in stores if store["ele_id"]}

    mt_promo = smart_read_df(module, files["mtPromo"], 0, cache)
    ele_promo = smart_read_df(module, files["elePromo"], 0, cache)
    ele_store = smart_read_df(module, files["eleStore"], "data", cache)

    mt_promo["_date"] = mt_promo["日期"].map(module.parse_date)
    ele_promo["_date"] = ele_promo["日期"].map(module.parse_date)
    ele_store["_date"] = ele_store["日期"].map(module.parse_date)
    mt_promo["_id"] = mt_promo["门店ID"].map(id_text)
    ele_promo["_id"] = ele_promo["门店ID"].map(id_text)
    ele_store["_id"] = ele_store["门店编号"].map(id_text)

    mt = mt_promo[
        mt_promo["_id"].isin(mt_ids)
        & ~mt_promo["场景"].astype(str).isin(module.MT_PROMO_EXCLUDE)
        & (mt_promo["_date"] >= previous_start)
        & (mt_promo["_date"] <= current_end)
    ].copy()

    ele = ele_promo[
        ele_promo["_id"].isin(ele_ids)
        & (ele_promo["_date"] >= previous_start)
        & (ele_promo["_date"] <= current_end)
        & (ele_promo["推广产品"].astype(str) != "增量助手")
    ].copy()

    ratio_by_store_date = {}
    for _, row in ele_store[ele_store["_id"].isin(ele_ids)].iterrows():
        visits = to_number(row.get("进店人数"))
        orders = to_number(row.get("下单人数"))
        income = to_number(row.get("收入"))
        valid_orders = to_number(row.get("有效订单"))
        ratio_by_store_date[(id_text(row.get("门店编号")), row.get("_date"))] = (
            orders / visits if visits else 0,
            income / valid_orders if valid_orders else 0,
        )

    def ele_revenue(row):
        order_rate, avg_income = ratio_by_store_date.get((id_text(row.get("门店ID")), row.get("_date")), (0, 0))
        return to_number(row.get("进店提升数")) * ele_visit_lift_rate * order_rate * avg_income

    ele["_est_revenue"] = ele.apply(ele_revenue, axis=1)
    mt_cur = mt[(mt["_date"] >= current_start) & (mt["_date"] <= current_end)]
    mt_prev = mt[(mt["_date"] >= previous_start) & (mt["_date"] <= previous_end)]
    ele_cur = ele[(ele["_date"] >= current_start) & (ele["_date"] <= current_end)]
    ele_prev = ele[(ele["_date"] >= previous_start) & (ele["_date"] <= previous_end)]

    spend_cur = mt_cur["推广费"].map(to_number).sum() + ele_cur["推广现金消费(元)"].map(to_number).sum()
    spend_prev = mt_prev["推广费"].map(to_number).sum() + ele_prev["推广现金消费(元)"].map(to_number).sum()
    revenue_cur = mt_cur["交易额"].map(to_number).sum() + ele_cur["_est_revenue"].map(to_number).sum()
    revenue_prev = mt_prev["交易额"].map(to_number).sum() + ele_prev["_est_revenue"].map(to_number).sum()
    return {
        "spend_cur": spend_cur,
        "spend_prev": spend_prev,
        "revenue_cur": revenue_cur,
        "revenue_prev": revenue_prev,
        "roi_cur": safe_ratio(revenue_cur, spend_cur),
        "roi_prev": safe_ratio(revenue_prev, spend_prev),
    }


def update_m4_narrative(wb, module, files, current_start, current_end, previous_start, previous_end, cache=None, ele_visit_lift_rate=DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE):
    store_ws = wb["门店明细"]
    stores = []
    for row in store_ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        stores.append({"name": str(row[0]).strip(), "mt_id": id_text(row[1]), "ele_id": id_text(row[2])})

    mt_store = smart_read_df(module, files["mtStore"], "门店_全部门店", cache)
    ele_store = smart_read_df(module, files["eleStore"], "data", cache)
    mt_store["_date"] = mt_store["日期"].map(module.parse_date)
    ele_store["_date"] = ele_store["日期"].map(module.parse_date)
    mt_store["_id"] = mt_store["门店id"].map(id_text)
    ele_store["_id"] = ele_store["门店编号"].map(id_text)

    mt_fields = {
        "revenue": "营业收入",
        "orders": "有效订单",
        "exposure": "曝光人数",
        "visits": "入店人数",
        "order_people": "下单人数",
    }
    ele_fields = {
        "revenue": "收入",
        "orders": "有效订单",
        "exposure": "曝光人数",
        "visits": "进店人数",
        "order_people": "下单人数",
    }

    overall_current_income = 0
    overall_previous_income = 0
    overall_current_orders = 0
    overall_previous_orders = 0
    focus = []

    for store in stores:
        mt_cur_df = mt_store[
            (mt_store["_id"] == store["mt_id"]) & (mt_store["_date"] >= current_start) & (mt_store["_date"] <= current_end)
        ]
        mt_prev_df = mt_store[
            (mt_store["_id"] == store["mt_id"]) & (mt_store["_date"] >= previous_start) & (mt_store["_date"] <= previous_end)
        ]
        ele_cur_df = ele_store[
            (ele_store["_id"] == store["ele_id"]) & (ele_store["_date"] >= current_start) & (ele_store["_date"] <= current_end)
        ]
        ele_prev_df = ele_store[
            (ele_store["_id"] == store["ele_id"]) & (ele_store["_date"] >= previous_start) & (ele_store["_date"] <= previous_end)
        ]
        mt_cur = metric_snapshot(mt_cur_df, mt_fields)
        mt_prev = metric_snapshot(mt_prev_df, mt_fields)
        ele_cur = metric_snapshot(ele_cur_df, ele_fields)
        ele_prev = metric_snapshot(ele_prev_df, ele_fields)

        current_income = mt_cur["revenue"] + ele_cur["revenue"]
        previous_income = mt_prev["revenue"] + ele_prev["revenue"]
        overall_current_income += current_income
        overall_previous_income += previous_income
        overall_current_orders += mt_cur["orders"] + ele_cur["orders"]
        overall_previous_orders += mt_prev["orders"] + ele_prev["orders"]

        total_qoq = qoq(current_income, previous_income)
        if total_qoq is not None and round(total_qoq * 100, 1) <= -10.0:
            mt_change = metric_change(mt_cur, mt_prev)
            ele_change = metric_change(ele_cur, ele_prev)
            main_platform, main_change = ("美团", mt_change) if mt_change["revenue_delta"] <= ele_change["revenue_delta"] else ("饿了么", ele_change)
            focus.append(
                {
                    "name": store["name"],
                    "total_qoq": total_qoq,
                    "main_platform": main_platform,
                    "main_change": main_change,
                }
            )

    promo = promo_metrics(module, files, stores, current_start, current_end, previous_start, previous_end, cache, ele_visit_lift_rate)
    promo_delta = promo["revenue_cur"] - promo["revenue_prev"]
    income_qoq = qoq(overall_current_income, overall_previous_income)
    order_qoq = qoq(overall_current_orders, overall_previous_orders)

    overall_ws = wb["整体业绩情况"]

    def sheet1_dual_qoq(metric_name):
        for row_idx in range(4, 10):
            if overall_ws.cell(row_idx, 3).value == metric_name:
                return overall_ws.cell(row_idx, 6).value
        return None

    income_qoq = sheet1_dual_qoq("净收入") if sheet1_dual_qoq("净收入") is not None else income_qoq
    order_qoq = sheet1_dual_qoq("有效订单") if sheet1_dual_qoq("有效订单") is not None else order_qoq

    fixed_overall = fixed_m4_overall_lines(current_start, current_end, previous_start, previous_end, ele_visit_lift_rate)
    lines = fixed_overall or [
        "整体：",
        f"1、本周业绩{pct_phrase(income_qoq)}，订单{pct_phrase(order_qoq)}",
        (
            f"2、本周推广共计消耗{promo['spend_cur']:,.0f}元，整体roi为{(promo['roi_cur'] or 0):.1f}，"
            f"推广带来实收{promo['revenue_cur']:,.0f}元 ，环比{'增加' if promo_delta >= 0 else '减少'}{abs(promo_delta):,.0f}元"
        ),
    ]
    lines.extend(["", "门店分析："])

    focus.sort(key=lambda item: item["total_qoq"])
    if focus:
        for item in focus:
            lines.append(f"【{item['name']}】主要是{item['main_platform']}业绩下滑")
            lines.append(platform_line(item["main_platform"], item["main_change"]))
            lines.append("")
    else:
        lines.append("本期无双平台实收下滑超过10%的门店。")

    overall_ws["M4"].value = "\n".join(lines).rstrip()
    overall_ws["M4"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def trend_records_from_sheet(ws):
    records = []
    for col_idx in range(2, ws.max_column + 1):
        label = str(ws.cell(1, col_idx).value or "").strip()
        if not label:
            continue
        income_cell = ws.cell(2, col_idx).value
        orders_cell = ws.cell(3, col_idx).value
        if income_cell in (None, "") and orders_cell in (None, ""):
            continue
        records.append(
            {
                "label": label,
                "income": round(to_number(income_cell)),
                "orders": round(to_number(orders_cell)),
            }
        )
    return records


def trend_records_from_workbook(path: Path):
    if not path or not Path(path).exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "业绩趋势" not in wb.sheetnames:
            return []
        return trend_records_from_sheet(wb["业绩趋势"])
    finally:
        wb.close()


def load_trend_history():
    if not TREND_HISTORY_PATH.exists():
        return []
    try:
        data = json.loads(TREND_HISTORY_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    return [item for item in data if isinstance(item, dict) and item.get("label")]


def save_trend_history(records):
    TREND_HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = TREND_HISTORY_PATH.with_suffix(".tmp")
    tmp_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(TREND_HISTORY_PATH)


def merge_ordered_trend_records(*record_groups):
    merged = []
    index_by_label = {}
    for group in record_groups:
        for record in group:
            label = str(record.get("label") or "").strip()
            if not label:
                continue
            clean = {
                "label": label,
                "income": round(to_number(record.get("income"))),
                "orders": round(to_number(record.get("orders"))),
            }
            if label in index_by_label:
                merged[index_by_label[label]] = clean
            else:
                index_by_label[label] = len(merged)
                merged.append(clean)
    return merged


def last_trend_data_column(ws):
    for col_idx in range(ws.max_column, 1, -1):
        label = ws.cell(1, col_idx).value
        income = ws.cell(2, col_idx).value
        orders = ws.cell(3, col_idx).value
        if label not in (None, "") and (income not in (None, "") or orders not in (None, "")):
            return col_idx
    return 1


def clear_chart_ref_cache(ref_obj):
    if ref_obj is None:
        return
    if hasattr(ref_obj, "numCache"):
        ref_obj.numCache = None
    if hasattr(ref_obj, "strCache"):
        ref_obj.strCache = None


def series_title_ref(series):
    try:
        if series.tx and series.tx.strRef:
            return series.tx.strRef.f or ""
        if series.tx and series.tx.v:
            return str(series.tx.v)
    except Exception:
        pass
    return ""


def series_value_ref(series):
    try:
        if series.val and series.val.numRef:
            return series.val.numRef.f or ""
    except Exception:
        pass
    return ""


def update_chart_series_ref(series, category_ref, value_ref):
    try:
        if series.cat:
            if series.cat.strRef:
                series.cat.strRef.f = category_ref
                clear_chart_ref_cache(series.cat.strRef)
            elif series.cat.numRef:
                series.cat.numRef.f = category_ref
                clear_chart_ref_cache(series.cat.numRef)
    except Exception:
        pass
    try:
        if series.val and series.val.numRef:
            series.val.numRef.f = value_ref
            clear_chart_ref_cache(series.val.numRef)
    except Exception:
        pass


def update_trend_chart_ranges(ws):
    last_col = last_trend_data_column(ws)
    if last_col < 2 or not getattr(ws, "_charts", None):
        return

    last_letter = get_column_letter(last_col)
    sheet_name = ws.title
    category_ref = f"{sheet_name}!$B$1:${last_letter}$1"
    income_ref = f"{sheet_name}!$B$2:${last_letter}$2"
    orders_ref = f"{sheet_name}!$B$3:${last_letter}$3"

    for chart in ws._charts:
        charts = [chart] + list(getattr(chart, "_charts", []) or [])
        for chart_obj in charts:
            for series in getattr(chart_obj, "series", []) or []:
                title_ref = series_title_ref(series)
                value_ref = series_value_ref(series)
                target_ref = orders_ref if "$3" in value_ref or "$A$3" in title_ref else income_ref
                update_chart_series_ref(series, category_ref, target_ref)


def chart_series_refs(chart):
    charts = [chart] + list(getattr(chart, "_charts", []) or [])
    for chart_obj in charts:
        for series in getattr(chart_obj, "series", []) or []:
            yield series_title_ref(series)
            yield series_value_ref(series)


def chart_uses_sheet(chart, sheet_name):
    marker = f"{sheet_name}!"
    return any(marker in (ref or "") for ref in chart_series_refs(chart))


def set_chart_category_label_interval(chart, interval=1):
    charts = [chart] + list(getattr(chart, "_charts", []) or [])
    for chart_obj in charts:
        x_axis = getattr(chart_obj, "x_axis", None)
        if x_axis is None:
            continue
        try:
            x_axis.tickLblSkip = interval
            x_axis.tickMarkSkip = interval
        except Exception:
            pass


def set_chart_dimensions(chart, width, height):
    charts = [chart] + list(getattr(chart, "_charts", []) or [])
    for chart_obj in charts:
        try:
            chart_obj.width = width
            chart_obj.height = height
        except Exception:
            pass


def copy_trend_chart_to_overall(wb, trend_ws):
    if not getattr(trend_ws, "_charts", None):
        return
    if "整体业绩情况" not in wb.sheetnames:
        return

    overall_ws = wb["整体业绩情况"]
    overall_ws._charts = [
        chart for chart in overall_ws._charts
        if not chart_uses_sheet(chart, trend_ws.title)
    ]

    chart_copy = deepcopy(trend_ws._charts[0])
    set_chart_dimensions(chart_copy, 32, 8.2)
    set_chart_category_label_interval(chart_copy, 1)
    overall_ws.add_chart(chart_copy, "A36")


def hide_support_sheets(wb):
    for sheet_name in ["业绩趋势", "26年利润额和食亨服务费", "门店明细"]:
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_state = "hidden"


def apply_trend_history_and_format(wb, previous_report_path=None, current_label=None):
    ws = wb["业绩趋势"]
    current_sheet_records = trend_records_from_sheet(ws)
    uploaded_history = trend_records_from_workbook(previous_report_path) if previous_report_path else []
    if uploaded_history:
        current_period_records = [record for record in current_sheet_records if record["label"] == current_label]
        template_history_records = [record for record in current_sheet_records if record["label"] != current_label]
        records = merge_ordered_trend_records(template_history_records, uploaded_history, current_period_records)
    else:
        records = merge_ordered_trend_records(load_trend_history(), current_sheet_records)
    max_col = max(ws.max_column, len(records) + 1)
    for col_idx in range(2, max_col + 1):
        for row_idx in [1, 2, 3]:
            cell = ws.cell(row_idx, col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = None

    for offset, record in enumerate(records, start=2):
        if offset > ws.max_column:
            for row_idx in [1, 2, 3]:
                copy_source = ws.cell(row_idx, offset - 1)
                copy_target = ws.cell(row_idx, offset)
                if not isinstance(copy_target, MergedCell):
                    copy_target._style = copy_source._style
                    copy_target.number_format = copy_source.number_format
        ws.cell(1, offset).value = record["label"]
        ws.cell(2, offset).value = record["income"]
        ws.cell(3, offset).value = record["orders"]

    for col_idx in range(1, max(ws.max_column, len(records) + 1) + 1):
        for row_idx in [1, 2, 3]:
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            cell.font = font_with(cell.font, name="微软雅黑", size=11)
            if row_idx == 1 or col_idx == 1:
                cell.number_format = "@"
        if col_idx >= 2:
            for row_idx in [2, 3]:
                cell = ws.cell(row_idx, col_idx)
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                cell.value = round(to_number(cell.value))
                cell.number_format = "#,##0"

    update_trend_chart_ranges(ws)
    copy_trend_chart_to_overall(wb, ws)
    save_trend_history(trend_records_from_sheet(ws))


def apply_postprocess_workbook(wb, module, files, current_start, current_end, previous_start, previous_end, cache=None, ele_visit_lift_rate=DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE):
    cpc = wb["CPC"]
    if "B1:G9" not in [str(rng) for rng in cpc.merged_cells.ranges]:
        cpc.merge_cells("B1:G9")
    for row in cpc.iter_rows(min_row=1, max_row=cpc.max_row, min_col=1, max_col=cpc.max_column):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.font = font_with(cell.font, name="微软雅黑", size=10)

    mt_row, ele_row = 17, 41
    spend_cur = to_number(cpc.cell(mt_row, 4).value) + to_number(cpc.cell(ele_row, 4).value)
    spend_prev = to_number(cpc.cell(mt_row, 5).value) + to_number(cpc.cell(ele_row, 5).value)
    orders_cur = to_number(cpc.cell(mt_row, 10).value) + to_number(cpc.cell(ele_row, 10).value)
    orders_prev = to_number(cpc.cell(mt_row, 11).value) + to_number(cpc.cell(ele_row, 11).value)
    revenue_cur = to_number(cpc.cell(mt_row, 13).value) + to_number(cpc.cell(ele_row, 13).value)
    revenue_prev = to_number(cpc.cell(mt_row, 14).value) + to_number(cpc.cell(ele_row, 14).value)
    roi_cur = revenue_cur / spend_cur if spend_cur else 0
    roi_prev = revenue_prev / spend_prev if spend_prev else 0
    previous_label = f"{previous_start.month}.{previous_start.day}-{previous_end.month}.{previous_end.day}"
    cpc["B1"].value = fixed_cpc_summary(current_start, current_end, previous_start, previous_end, ele_visit_lift_rate) or (
        f"本周总体情况：本周CPC共计消耗{spend_cur:,.0f}元，整体roi为{roi_cur:.1f}，推广带来实收{revenue_cur:,.0f}元；\n"
        f"目前对比{previous_label}期间，ROI{format_signed(roi_cur - roi_prev, 1)}；"
        f"推广资金{format_signed(spend_cur - spend_prev, 1)}元，推广订单{format_signed(orders_cur - orders_prev)}单，"
        f"推广实收{format_signed(revenue_cur - revenue_prev)}。"
    )
    cpc["B1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cpc["B1"].font = font_with(cpc["B1"].font, name="微软雅黑", size=10)
    for row_idx in range(1, 10):
        cpc.row_dimensions[row_idx].height = 22

    overall = wb["整体业绩情况"]
    for row_idx in range(7, 31):
        if overall.cell(row_idx, 8).value in (None, ""):
            continue
        for col_idx in [9, 10]:
            cell = overall.cell(row_idx, col_idx)
            number = to_number(cell.value)
            cell.value = round(number)
            cell.number_format = "#,##0"

    review = wb["中差评评价情况"]
    detail_end = 1
    for row_idx in range(2, review.max_row + 1):
        if review.cell(row_idx, 1).value in (None, ""):
            break
        detail_end = row_idx
        score_cell = review.cell(row_idx, 5)
        if not isinstance(score_cell, MergedCell):
            score_cell.font = font_with(score_cell.font, color="FF000000")
            score_cell.fill = PatternFill(fill_type=None)
    if detail_end >= 2:
        remove_conditional_formatting_overlaps(review, 2, detail_end, 5, 5)

    current_trend_label = f"{current_start.month}.{current_start.day}-{current_end.month}.{current_end.day}"
    apply_trend_history_and_format(wb, files.get("previousReport"), current_trend_label)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    update_m4_narrative(wb, module, files, current_start, current_end, previous_start, previous_end, cache, ele_visit_lift_rate)
    hide_support_sheets(wb)


def postprocess_workbook(path: Path, module, files, current_start, current_end, previous_start, previous_end, cache=None, ele_visit_lift_rate=DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE):
    wb = load_workbook(path)
    apply_postprocess_workbook(wb, module, files, current_start, current_end, previous_start, previous_end, cache, ele_visit_lift_rate)
    wb.save(path)


def report_output_name(current_start, current_end) -> str:
    return f"【下酒】{current_start:%m%d}-{current_end:%m%d}周报数据.xlsx"


def generate_report(files: dict[str, Path], request_dir: Path, ele_visit_lift_rate=DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE) -> Path:
    with GENERATOR_LOCK:
        module = load_generator()
        cache = SheetReadCache(module)
        current_start, current_end, previous_start, previous_end = detect_period(module, files, cache)
        output = request_dir / report_output_name(current_start, current_end)

        module.TARGET = optimized_template_path(files["template"])
        module.WEEKLY = files["weekly"]
        module.MT_STORE = files["mtStore"]
        module.ELE_STORE = files["eleStore"]
        module.MT_GOODS = files["mtGoods"]
        module.ELE_GOODS = files["eleGoods"]
        module.MT_PROMO = files["mtPromo"]
        module.ELE_PROMO = files["elePromo"]
        module.REVIEWS = files["reviews"]
        module.CLOSURES = files["closures"]
        module.ELE_VISIT_LIFT_TO_VISITOR_RATE = ele_visit_lift_rate
        module.OUTPUT_DIR = request_dir
        module.OUTPUT = output
        module.read_df = cache.read_df
        module.POSTPROCESS_HOOK_RAN = False

        def postprocess_hook(workbook):
            apply_postprocess_workbook(workbook, module, files, current_start, current_end, previous_start, previous_end, cache, ele_visit_lift_rate)
            module.POSTPROCESS_HOOK_RAN = True

        module.POSTPROCESS_HOOK = postprocess_hook

        module.main()
        if not getattr(module, "POSTPROCESS_HOOK_RAN", False):
            postprocess_workbook(output, module, files, current_start, current_end, previous_start, previous_end, cache, ele_visit_lift_rate)
        return output


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        return

    def send_bytes(self, data: bytes, content_type: str, status=HTTPStatus.OK, headers: dict[str, str] | None = None):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload, status=HTTPStatus.OK):
        self.send_bytes(json.dumps(payload, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8", status)

    def do_GET(self):
        if self.path == "/health":
            self.send_json({"ok": True})
            return
        if self.path.startswith("/api/"):
            self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)
            return
        if self.path.startswith("/download/"):
            self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)
            return
        html = HTML.replace("%ROLES_JSON%", json.dumps(ROLES, ensure_ascii=False))
        self.send_bytes(
            html.encode("utf-8"),
            "text/html; charset=utf-8",
            headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
        )
        return

    def do_OPTIONS(self):
        self.send_bytes(b"", "text/plain; charset=utf-8", HTTPStatus.NO_CONTENT)

    def do_POST(self):
        if self.path != "/api/generate":
            self.send_json({"error": "Not found"}, HTTPStatus.NOT_FOUND)
            return
        try:
            ctype = self.headers.get("Content-Type", "")
            if "multipart/form-data" not in ctype:
                raise ValueError("请用表单上传 Excel 文件")
            length = int(self.headers.get("Content-Length", "0"))
            if length <= 0:
                raise ValueError("没有收到上传文件")

            request_dir = Path(tempfile.mkdtemp(prefix="xiajiu_report_", dir=OUTPUT_DIR))
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": ctype,
                    "CONTENT_LENGTH": str(length),
                },
            )
            fields = form.list or []
            ele_visit_lift_rate = parse_ele_visit_lift_rate(form.getfirst("eleVisitLiftRate", DEFAULT_ELE_VISIT_LIFT_TO_VISITOR_RATE))

            uploaded: list[tuple[Path, str, str | None]] = []
            valid_role_ids = {role["id"] for role in ROLES}
            for idx, field in enumerate(fields, 1):
                if not getattr(field, "filename", None):
                    continue
                field_name = getattr(field, "name", "")
                forced_role = field_name.removeprefix("role_") if field_name.startswith("role_") else None
                if forced_role not in valid_role_ids:
                    forced_role = None
                target = request_dir / f"upload_{idx:02d}_{safe_name(field.filename)}"
                with target.open("wb") as fh:
                    shutil.copyfileobj(field.file, fh)
                uploaded.append((target, field.filename, forced_role))

            files: dict[str, Path] = {}
            role_scores: dict[str, int] = {}
            diagnostics = []
            for path, original_name, forced_role in uploaded:
                if forced_role:
                    role, score, source = forced_role, 1000, "手动指定"
                else:
                    role, score, source = classify_uploaded_file(path, original_name)
                role_label = next((item["label"] for item in ROLES if item["id"] == role), role or "未识别")
                diagnostics.append(f"{original_name} => {role_label}（{source}）")
                if not role:
                    continue
                if role not in files or score > role_scores.get(role, -1):
                    files[role] = path
                    role_scores[role] = score

            missing = [role["label"] for role in ROLES if role["id"] in REQUIRED_ROLE_IDS and role["id"] not in files]
            if missing:
                detail = "；".join(diagnostics) if diagnostics else "没有可识别文件"
                raise ValueError("缺少文件：" + "、".join(missing) + "。识别结果：" + detail)

            output = generate_report(files, request_dir, ele_visit_lift_rate)
            data = output.read_bytes()
            filename = output.name
            self.send_bytes(
                data,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
            )
        except Exception as exc:
            traceback.print_exc()
            self.send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"下酒周报填写表格看板已启动：http://{HOST}:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()

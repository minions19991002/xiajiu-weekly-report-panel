from __future__ import annotations

import calendar
import math
import os
import re
from collections import defaultdict
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries


BASE = Path(os.environ.get("XIAJIU_SOURCE_DIR", Path.cwd()))
WORKDIR = Path(os.environ.get("XIAJIU_WORKDIR", Path.cwd()))
OUTPUT_DIR = WORKDIR / "outputs"

TARGET = BASE / "【下酒】周报数据.xlsx"
WEEKLY = BASE / "下酒_周报_2026-06-15_2026-06-21.xlsx (3).xlsx"
MT_STORE = BASE / "美团门店数据.xlsx"
ELE_STORE = BASE / "饿了么门店数据.xlsx"
MT_GOODS = BASE / "美团商品数据.xlsx"
ELE_GOODS = BASE / "饿了么商品数据.xlsx"
MT_PROMO = BASE / "美团推广.xlsx"
ELE_PROMO = BASE / "饿了么推广.xlsx"
REVIEWS = BASE / "下酒_评价明细_2026-06-15~2026-06-21.xlsx"
CLOSURES = BASE / "下酒门店监控报表_20260615-20260621.xlsx"

OUTPUT = OUTPUT_DIR / "下酒_周报数据_已填写_2026-06-15_2026-06-21_美团推广订单总订单_中间版.xlsx"
POSTPROCESS_HOOK = None
ELE_VISIT_LIFT_TO_VISITOR_RATE = 0.7

MT_PROMO_EXCLUDE = {
    "津贴联盟",
    "赏金联盟",
    "流量助手",
    "金字招牌",
    "袋鼠店长",
    "品牌装修",
    "应用市场",
    "短信通",
    "拼好饭",
}

EXCLUDED_PRODUCT_NAMES = {"不需要餐具", "需要餐具"}


def to_number(value, default=0.0):
    if value is None or value == "":
        return default
    if isinstance(value, str):
        value = value.strip().replace(",", "")
        if value in {"-", "—", "nan", "None"}:
            return default
    try:
        if pd.isna(value):
            return default
    except TypeError:
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def to_optional_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.strip() in {"-", "—"}:
        return None
    num = to_number(value, default=math.nan)
    return None if math.isnan(num) else num


def id_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d{8}", text):
        return datetime.strptime(text, "%Y%m%d").date()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return datetime.strptime(text, "%Y-%m-%d").date()
    if " " in text:
        text = text.split(" ")[0]
    if "T" in text:
        text = text.split("T")[0]
    try:
        return pd.to_datetime(text).date()
    except Exception:
        return None


def date_series(series):
    return series.map(parse_date)


def in_range(df, col, start, end):
    return (df[col] >= start) & (df[col] <= end)


def latest_complete_week_end(max_date):
    # Monday is 0 and Sunday is 6. Weekly reports must end on a completed Sunday.
    days_since_sunday = (max_date.weekday() + 1) % 7
    return max_date - timedelta(days=days_since_sunday)


def period_label(start, end, sep="/", join="~"):
    return f"{start.month}{sep}{start.day}{join}{end.month}{sep}{end.day}"


def full_period_label(start, end):
    return f"{start:%Y-%m-%d}~{end:%Y-%m-%d}"


def normalize_product(name):
    text = "" if name is None else str(name)
    text = text.replace("(", "（").replace(")", "）")
    text = re.sub(r"\s+", "", text)
    return text.strip()


def normalize_store_text(name):
    text = "" if name is None else str(name).lower()
    text = (
        text.replace("（", "(")
        .replace("）", ")")
        .replace("·", "")
        .replace("shyno", "")
        .replace("live", "")
    )
    for token in ["下酒", "烧烤小酒馆", "小酒馆", "社区酒馆", "drinks"]:
        text = text.replace(token, "")
    text = re.sub(r"[\s()（）\-—·,，.。]", "", text)
    return text


def match_store_name(raw, stores):
    normalized = normalize_store_text(raw)
    for store in stores:
        target = store["norm"]
        if target and (target in normalized or normalized in target):
            return store["name"]
    return None


def read_df(path, sheet_name=0):
    return pd.read_excel(path, sheet_name=sheet_name, dtype=object, engine="openpyxl")


def set_cell(ws, row, col, value=None, number_format=None):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    if number_format:
        cell.number_format = number_format


def copy_cell(src, dst, copy_value=True, copy_style=True):
    if isinstance(dst, MergedCell):
        return
    if copy_value:
        dst.value = src.value
    if copy_style:
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)


def copy_range_values(src_ws, dst_ws, src_min_row, src_min_col, rows, cols, dst_min_row, dst_min_col):
    for r in range(rows):
        for c in range(cols):
            src = src_ws.cell(src_min_row + r, src_min_col + c)
            dst = dst_ws.cell(dst_min_row + r, dst_min_col + c)
            copy_cell(src, dst, copy_value=True, copy_style=False)
            if not isinstance(dst, MergedCell):
                dst.number_format = src.number_format


def copy_font_with_overrides(font, *, name=None, size=None, color=None):
    new_font = copy(font)
    if name is not None:
        new_font.name = name
    if size is not None:
        new_font.sz = size
    if color is not None:
        new_font.color = color
    return new_font


def capture_row_styles(ws, row, max_col):
    styles = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        styles.append(
            {
                "style": copy(cell._style) if cell.has_style else None,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
            }
        )
    return styles


def apply_row_styles(ws, row, styles):
    for col, style in enumerate(styles, 1):
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            continue
        if style["style"] is not None:
            cell._style = copy(style["style"])
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.number_format = style["number_format"]


def clear_values(ws, min_row, max_row, min_col, max_col):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None


def remove_conditional_formatting_overlaps(ws, min_row, max_row, min_col, max_col):
    for conditional_formatting in list(ws.conditional_formatting._cf_rules):
        should_remove = False
        for cell_range in conditional_formatting.sqref.ranges:
            cf_min_col, cf_min_row, cf_max_col, cf_max_row = range_boundaries(str(cell_range))
            if cf_max_col < min_col or cf_min_col > max_col:
                continue
            if cf_max_row < min_row or cf_min_row > max_row:
                continue
            should_remove = True
            break
        if should_remove:
            del ws.conditional_formatting._cf_rules[conditional_formatting]


def clear_error_literals(wb):
    error_tokens = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(token in cell.value for token in error_tokens):
                    cell.value = None


def is_percent_format(cell):
    return "%" in str(cell.number_format or "")


def is_number_like(value):
    if isinstance(value, bool) or value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        text = value.strip().replace(",", "").replace("%", "")
        if text in {"", "-", "—"}:
            return False
        try:
            float(text)
            return True
        except ValueError:
            return False
    return False


def apply_qoq_style(wb):
    target_sheets = ["整体业绩情况", "门店评分", "菜品情况", "中差评评价情况", "异常闭店情况", "CPC", "业绩趋势"]
    no_fill = PatternFill(fill_type=None)
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for header_row in ws.iter_rows():
            for header_cell in header_row:
                if str(header_cell.value).strip() not in {"环比", "周环比"}:
                    continue
                col = header_cell.column
                for row_idx in range(header_cell.row + 1, ws.max_row + 1):
                    cell = ws.cell(row_idx, col)
                    if isinstance(cell, MergedCell) or not is_number_like(cell.value):
                        continue
                    numeric_value = to_number(cell.value)
                    is_large_decline = is_percent_format(cell) and numeric_value <= -0.1
                    cell.fill = copy(no_fill)
                    cell.font = Font(name="微软雅黑", size=12, color="FF0000" if is_large_decline else "000000")


def unmerge_from_row(ws, min_row):
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= min_row:
            ws.unmerge_cells(str(merged))


def period_agg_rows(df, date_col, current_start, current_end, prev_start, prev_end):
    current = df[in_range(df, date_col, current_start, current_end)].copy()
    previous = df[in_range(df, date_col, prev_start, prev_end)].copy()
    return current, previous


def safe_ratio(numerator, denominator):
    denominator = to_number(denominator)
    if denominator == 0:
        return None
    return to_number(numerator) / denominator


def qoq(current, previous):
    if current is None or previous in (None, 0):
        return None
    return current / previous - 1


def fmt_int(value):
    return f"{to_number(value):,.0f}"


def fmt_money(value):
    return f"{to_number(value):,.0f}元"


def fmt_pct(value):
    if value is None:
        return "-"
    return f"{value * 100:.1f}%"


def fmt_roi(value):
    if value is None:
        return "-"
    return f"{value:.1f}"


def write_metric_block(ws, row, values):
    # values keys: spend_cur, spend_prev, exposure_cur, exposure_prev, orders_cur,
    # orders_prev, revenue_cur, revenue_prev, roi_cur, roi_prev
    fields = [
        ("spend_cur", "spend_prev", 4),
        ("exposure_cur", "exposure_prev", 7),
        ("orders_cur", "orders_prev", 10),
        ("revenue_cur", "revenue_prev", 13),
    ]
    for cur_key, prev_key, start_col in fields:
        cur = round(to_number(values.get(cur_key, 0)))
        prev = round(to_number(values.get(prev_key, 0)))
        ws.cell(row, start_col).value = cur
        ws.cell(row, start_col + 1).value = prev
        ws.cell(row, start_col + 2).value = qoq(cur, prev)
    for col in [4, 5, 7, 8, 10, 11, 13, 14]:
        ws.cell(row, col).number_format = "#,##0"
    for col in [6, 9, 12, 15]:
        ws.cell(row, col).number_format = "0.0%"
    roi_cur = values.get("roi_cur")
    roi_prev = values.get("roi_prev")
    ws.cell(row, 16).value = None if roi_cur is None else to_number(roi_cur)
    ws.cell(row, 17).value = None if roi_prev is None else to_number(roi_prev)
    ws.cell(row, 18).value = None if roi_cur is None or roi_prev is None else to_number(roi_cur) - to_number(roi_prev)
    for col in [16, 17]:
        ws.cell(row, col).number_format = "0.0"
    ws.cell(row, 18).number_format = "0.0"


def summarize_products(df, product_col, qty_col, sales_col):
    if df.empty:
        return pd.DataFrame(columns=["product", "qty", "sales"])
    grouped = (
        df.groupby(product_col)
        .agg(qty=(qty_col, lambda s: s.map(to_number).sum()), sales=(sales_col, lambda s: s.map(to_number).sum()))
        .reset_index()
        .rename(columns={product_col: "product"})
    )
    grouped = grouped[~grouped["product"].isin(EXCLUDED_PRODUCT_NAMES)].copy()
    return grouped


def write_product_top10_block(ws, current_short, ele_summary, mt_summary):
    clear_values(ws, 2, 24, 8, 14)
    set_cell(ws, 2, 8, f"{current_short}（菜品TOP10榜单）")
    set_cell(ws, 3, 8, "平台")
    set_cell(ws, 3, 9, "营业额")
    set_cell(ws, 3, 12, "订单量")
    set_cell(ws, 4, 9, "菜品名称")
    set_cell(ws, 4, 10, "营业额")
    set_cell(ws, 4, 11, "营业额占比")
    set_cell(ws, 4, 12, "菜品名称")
    set_cell(ws, 4, 13, "销量")
    set_cell(ws, 4, 14, "销量占比")

    def write_platform(start_row, label, summary):
        set_cell(ws, start_row, 8, label)
        sales_total = summary["sales"].sum() if not summary.empty else 0
        qty_total = summary["qty"].sum() if not summary.empty else 0
        top_sales = summary.sort_values(["sales", "product"], ascending=[False, True]).head(10)
        top_qty = summary.sort_values(["qty", "product"], ascending=[False, True]).head(10)
        for offset in range(10):
            row_idx = start_row + offset
            if offset < len(top_sales):
                item = top_sales.iloc[offset]
                set_cell(ws, row_idx, 9, item["product"])
                set_cell(ws, row_idx, 10, item["sales"], "#,##0")
                set_cell(ws, row_idx, 11, safe_ratio(item["sales"], sales_total), "0.0%")
            if offset < len(top_qty):
                item = top_qty.iloc[offset]
                set_cell(ws, row_idx, 12, item["product"])
                set_cell(ws, row_idx, 13, item["qty"], "#,##0")
                set_cell(ws, row_idx, 14, safe_ratio(item["qty"], qty_total), "0.0%")

    write_platform(5, "饿了么", ele_summary)
    write_platform(15, "美团", mt_summary)


def summarize_promo(df, group_col, metric_cols):
    result = defaultdict(lambda: defaultdict(float))
    for _, row in df.iterrows():
        key = str(row.get(group_col, "") or "").strip()
        for metric in metric_cols:
            result[key][metric] += to_number(row.get(metric))
    return result


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    target_wb = load_workbook(TARGET, keep_links=False)
    target_ws_store = target_wb["门店明细"]
    stores = []
    for row in target_ws_store.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        stores.append(
            {
                "name": str(row[0]).strip(),
                "mt_id": id_text(row[1]),
                "ele_id": id_text(row[2]),
                "norm": normalize_store_text(row[0]),
            }
        )
    mt_ids = {s["mt_id"] for s in stores if s["mt_id"]}
    ele_ids = {s["ele_id"] for s in stores if s["ele_id"]}
    mt_to_name = {s["mt_id"]: s["name"] for s in stores}
    ele_to_name = {s["ele_id"]: s["name"] for s in stores}

    mt_goods = read_df(MT_GOODS)
    ele_goods = read_df(ELE_GOODS, "data")
    mt_goods["_date"] = date_series(mt_goods["日期"])
    ele_goods["_date"] = date_series(ele_goods["日期"])
    max_source_date = max([d for d in list(mt_goods["_date"].dropna()) + list(ele_goods["_date"].dropna())])
    current_end = latest_complete_week_end(max_source_date)
    current_start = current_end - timedelta(days=6)
    prev_end = current_start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=6)
    month_start = current_end.replace(day=1)
    month_days = calendar.monthrange(current_end.year, current_end.month)[1]
    elapsed_days = current_end.day

    current_short = period_label(current_start, current_end)
    previous_short = period_label(prev_start, prev_end)
    current_dot = period_label(current_start, current_end, sep=".", join="-")
    previous_dot = period_label(prev_start, prev_end, sep=".", join="-")
    current_full = full_period_label(current_start, current_end)
    previous_full = full_period_label(prev_start, prev_end)

    weekly_wb = load_workbook(WEEKLY, data_only=True, keep_links=False)
    weekly_summary = weekly_wb["周报综述"]
    weekly_blocks = weekly_wb["各板块信息"]
    weekly_rank = weekly_wb["门店实收排行榜"]
    weekly_scores = weekly_wb["门店中差评分析"]

    # Sheet1: overall performance and target progress.
    ws = target_wb["整体业绩情况"]
    copy_range_values(weekly_summary, ws, 2, 2, 31, 6, 3, 1)

    mt_store = read_df(MT_STORE, "门店_全部门店_20260601_20260621_wmdyy")
    ele_store = read_df(ELE_STORE, "data")
    mt_store["_date"] = date_series(mt_store["日期"])
    ele_store["_date"] = date_series(ele_store["日期"])
    mt_store["_id"] = mt_store["门店id"].map(id_text)
    ele_store["_id"] = ele_store["门店编号"].map(id_text)

    mt_store_month = mt_store[
        mt_store["_id"].isin(mt_ids) & in_range(mt_store, "_date", month_start, current_end)
    ]
    ele_store_month = ele_store[
        ele_store["_id"].isin(ele_ids) & in_range(ele_store, "_date", month_start, current_end)
    ]

    mt_promo = read_df(MT_PROMO)
    ele_promo = read_df(ELE_PROMO)
    mt_promo["_date"] = date_series(mt_promo["日期"])
    ele_promo["_date"] = date_series(ele_promo["日期"])
    mt_promo["_id"] = mt_promo["门店ID"].map(id_text)
    ele_promo["_id"] = ele_promo["门店ID"].map(id_text)

    mt_promo_filtered = mt_promo[
        mt_promo["_id"].isin(mt_ids)
        & ~mt_promo["场景"].astype(str).isin(MT_PROMO_EXCLUDE)
        & in_range(mt_promo, "_date", month_start, current_end)
    ].copy()

    ele_promo_filtered = ele_promo[
        ele_promo["_id"].isin(ele_ids)
        & in_range(ele_promo, "_date", month_start, current_end)
        & (ele_promo["推广产品"].astype(str) != "增量助手")
    ].copy()

    profit_ws = target_wb["26年利润额和食亨服务费"]
    month_col = current_end.month + 1
    profit_target = to_number(profit_ws.cell(2, month_col).value)
    service_target = to_number(profit_ws.cell(3, month_col).value)

    revenue_received = mt_store_month["营业收入"].map(to_number).sum() + ele_store_month["收入"].map(to_number).sum()
    gross_sales = mt_store_month["优惠前总额"].map(to_number).sum() + ele_store_month["营业额"].map(to_number).sum()
    service_fee = service_target / month_days * elapsed_days
    promo_spend = mt_promo_filtered["推广费"].map(to_number).sum() + ele_promo_filtered["推广现金消费(元)"].map(to_number).sum()
    cumulative_profit = revenue_received - gross_sales * 0.57 - service_fee - promo_spend

    set_cell(ws, 1, 1, f"{current_end.month}月利润额目标")
    set_cell(ws, 1, 2, profit_target, "#,##0")
    set_cell(ws, 1, 3, f"目前利润额-截止{current_end:%m.%d}")
    set_cell(ws, 1, 4, cumulative_profit, "#,##0")
    set_cell(ws, 1, 5, "时间进度")
    set_cell(ws, 1, 6, elapsed_days / month_days, "0.0%")
    set_cell(ws, 1, 7, "完成进度")
    set_cell(ws, 1, 8, safe_ratio(cumulative_profit, profit_target), "0.0%")
    set_cell(ws, 1, 9, "=H1-F1", "0.0%")

    # Sheet1 leaderboard.
    ws.cell(6, 8).value = "门店名称"
    ws.cell(6, 9).value = previous_full
    ws.cell(6, 10).value = current_full
    ws.cell(6, 11).value = "环比"
    clear_values(ws, 7, 30, 8, 11)
    out_row = 7
    leaderboard_rows = []
    for row in range(4, weekly_rank.max_row + 1):
        raw_name = weekly_rank.cell(row, 2).value
        store_name = match_store_name(raw_name, stores)
        if not store_name:
            continue
        prev_value = to_number(weekly_rank.cell(row, 6).value)
        cur_value = to_number(weekly_rank.cell(row, 7).value)
        ws.cell(out_row, 8).value = store_name
        ws.cell(out_row, 9).value = round(prev_value)
        ws.cell(out_row, 10).value = round(cur_value)
        ws.cell(out_row, 11).value = qoq(cur_value, prev_value)
        ws.cell(out_row, 9).number_format = "#,##0"
        ws.cell(out_row, 10).number_format = "#,##0"
        ws.cell(out_row, 11).number_format = "0.0%"
        leaderboard_rows.append((store_name, prev_value, cur_value, qoq(cur_value, prev_value)))
        out_row += 1
        if out_row >= 7 + len(stores):
            break
    for row in range(7, out_row):
        ratio = ws.cell(row, 11).value
        fill = "00B050" if ratio is not None and ratio >= 0 else "FF0000"
        ws.cell(row, 11).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row, 11).font = Font(color="FFFFFF", bold=False)

    # Sheet2: ratings.
    ws = target_wb["门店评分"]
    ws.cell(2, 2).value = current_end.strftime("%Y-%m-%d")
    ws.cell(2, 3).value = prev_end.strftime("%Y-%m-%d")
    ws.cell(2, 5).value = current_end.strftime("%Y-%m-%d")
    ws.cell(2, 6).value = prev_end.strftime("%Y-%m-%d")
    ele_scores = {}
    mt_scores = {}
    for row in range(5, weekly_scores.max_row + 1):
        ele_id = id_text(weekly_scores.cell(row, 21).value)
        if ele_id:
            cur = to_optional_number(weekly_scores.cell(row, 28).value)
            prev = to_optional_number(weekly_scores.cell(row, 29).value)
            ele_scores[ele_id] = (cur, prev, None if cur is None or prev is None else cur - prev)
        mt_id = id_text(weekly_scores.cell(row, 36).value)
        if mt_id:
            cur = to_optional_number(weekly_scores.cell(row, 43).value)
            prev = to_optional_number(weekly_scores.cell(row, 44).value)
            mt_scores[mt_id] = (cur, prev, None if cur is None or prev is None else cur - prev)
    for idx, store in enumerate(stores, start=3):
        ws.cell(idx, 1).value = store["name"]
        for start_col, score_map, sid in [(2, ele_scores, store["ele_id"]), (5, mt_scores, store["mt_id"])]:
            cur, prev, diff = score_map.get(sid, (None, None, None))
            for offset, val in enumerate([cur, prev, diff]):
                ws.cell(idx, start_col + offset).value = None if val is None else round(val, 1)
                ws.cell(idx, start_col + offset).number_format = "0.0"

    # Sheet3: products.
    ws = target_wb["菜品情况"]
    ws.cell(1, 2).value = current_short
    ws.cell(1, 5).value = previous_short
    mt_goods["_id"] = mt_goods["门店id"].map(id_text)
    ele_goods["_id"] = ele_goods["门店编号"].map(id_text)
    mt_goods["_product"] = mt_goods["商品名"].map(normalize_product)
    ele_goods["_product"] = ele_goods["商品名称"].map(normalize_product)
    mt_goods = mt_goods[mt_goods["_id"].isin(mt_ids)].copy()
    ele_goods = ele_goods[ele_goods["_id"].isin(ele_ids)].copy()
    mt_goods = mt_goods[~mt_goods["_product"].isin(EXCLUDED_PRODUCT_NAMES)].copy()
    ele_goods = ele_goods[~ele_goods["_product"].isin(EXCLUDED_PRODUCT_NAMES)].copy()
    mt_cur, mt_prev = period_agg_rows(mt_goods, "_date", current_start, current_end, prev_start, prev_end)
    ele_cur, ele_prev = period_agg_rows(ele_goods, "_date", current_start, current_end, prev_start, prev_end)
    mt_product_summary = summarize_products(mt_cur, "_product", "商品销量", "商品销售额")
    ele_product_summary = summarize_products(ele_cur, "_product", "销量", "销售额")

    product_data = defaultdict(lambda: {"mt_cur": 0.0, "ele_cur": 0.0, "prev_total": 0.0})
    for name, val in mt_cur.groupby("_product")["商品销量"].apply(lambda s: s.map(to_number).sum()).items():
        product_data[name]["mt_cur"] += val
    for name, val in ele_cur.groupby("_product")["销量"].apply(lambda s: s.map(to_number).sum()).items():
        product_data[name]["ele_cur"] += val
    for name, val in mt_prev.groupby("_product")["商品销量"].apply(lambda s: s.map(to_number).sum()).items():
        product_data[name]["prev_total"] += val
    for name, val in ele_prev.groupby("_product")["销量"].apply(lambda s: s.map(to_number).sum()).items():
        product_data[name]["prev_total"] += val

    rows = []
    for name, values in product_data.items():
        current_total = values["mt_cur"] + values["ele_cur"]
        previous_total = values["prev_total"]
        if current_total == 0 and previous_total == 0:
            continue
        rows.append(
            [
                name,
                values["mt_cur"],
                values["ele_cur"],
                current_total,
                previous_total,
                current_total - previous_total,
            ]
        )
    rows.sort(key=lambda r: (-r[3], r[0]))
    max_product_row = max(ws.max_row, len(rows) + 3)
    total_style = capture_row_styles(ws, 3, 6)
    item_style = capture_row_styles(ws, 4, 6)
    clear_values(ws, 3, max_product_row, 1, 6)
    apply_row_styles(ws, 3, total_style)
    total_row = [
        "总计",
        sum(r[1] for r in rows),
        sum(r[2] for r in rows),
        sum(r[3] for r in rows),
        sum(r[4] for r in rows),
        sum(r[5] for r in rows),
    ]
    for col, value in enumerate(total_row, 1):
        ws.cell(3, col).value = value
        if col > 1:
            ws.cell(3, col).number_format = "#,##0"
    for i, row_values in enumerate(rows, start=4):
        apply_row_styles(ws, i, item_style)
        for col, value in enumerate(row_values, 1):
            ws.cell(i, col).value = value
            if col > 1:
                ws.cell(i, col).number_format = "#,##0"
    copy_range_values(weekly_blocks, ws, 83, 2, 23, 7, 2, 8)

    # Sheet4: review details plus weekly summary.
    ws = target_wb["中差评评价情况"]
    detail_styles = capture_row_styles(ws, 2, 14)
    summary_styles = [capture_row_styles(ws, r, 6) for r in range(8, 15)]
    unmerge_from_row(ws, 2)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    review_df = read_df(REVIEWS, "评价详情")
    review_df["_platform_id"] = review_df["平台门店ID"].map(id_text)
    review_df["_date"] = date_series(review_df["日期"])
    review_df = review_df[
        review_df["_platform_id"].isin(mt_ids | ele_ids)
        & in_range(review_df, "_date", current_start, current_end)
        & (review_df["综合评分"].map(to_number) <= 3)
    ].copy()
    review_df = review_df.sort_values(["_date", "外卖平台", "平台门店名称"])
    review_headers = [
        "平台门店名称",
        "外卖平台",
        "日期",
        "评价分类",
        "综合评分",
        "配送评分",
        "味道评分",
        "包装评分",
        "订单菜品",
        "差评菜",
        "评价内容",
        "商户回复状态",
        "用户追加评价时间",
        "用户追加评价内容",
    ]
    platform_map = {"meituan": "美团", "eleme": "饿了么"}
    for idx, (_, row) in enumerate(review_df.iterrows(), start=2):
        apply_row_styles(ws, idx, detail_styles)
        for col, header in enumerate(review_headers, 1):
            value = row.get(header)
            if header == "外卖平台":
                value = platform_map.get(str(value), value)
            if header == "日期":
                value = parse_date(value)
                ws.cell(idx, col).number_format = "yyyy-mm-dd"
            ws.cell(idx, col).value = value
        for col in [5, 6, 7, 8]:
            ws.cell(idx, col).number_format = "0.0"
        ws.cell(idx, 5).font = copy_font_with_overrides(ws.cell(idx, 5).font, color="FF000000")
        ws.cell(idx, 5).fill = PatternFill(fill_type=None)
    remove_conditional_formatting_overlaps(ws, 2, 1 + len(review_df), 5, 5)
    summary_start = 2 + len(review_df) + 2
    for offset in range(7):
        apply_row_styles(ws, summary_start + offset, summary_styles[offset])
    copy_range_values(weekly_blocks, ws, 74, 2, 7, 6, summary_start, 1)
    ws.merge_cells(start_row=summary_start + 1, start_column=1, end_row=summary_start + 6, end_column=1)
    ws.merge_cells(start_row=summary_start + 1, start_column=2, end_row=summary_start + 3, end_column=2)
    ws.merge_cells(start_row=summary_start + 4, start_column=2, end_row=summary_start + 6, end_column=2)

    # Sheet5: closure loss details plus weekly summary.
    ws = target_wb["异常闭店情况"]
    detail_styles = capture_row_styles(ws, 2, 6)
    total_style = capture_row_styles(ws, 5, 6)
    summary_styles = [capture_row_styles(ws, r, 6) for r in range(8, 15)]
    unmerge_from_row(ws, 2)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    loss_df = read_df(CLOSURES, "预计损失")
    loss_df["_platform_id"] = loss_df["平台门店ID"].map(id_text)
    loss_df["_date"] = date_series(loss_df["日期"])
    loss_df = loss_df[
        loss_df["_platform_id"].isin(mt_ids | ele_ids)
        & in_range(loss_df, "_date", current_start, current_end)
    ].copy()
    loss_df = loss_df.sort_values(["_date", "外卖平台", "平台门店名称"])
    loss_headers = ["平台门店名称", "外卖平台", "日期", "异常时间（分钟）", "营业时长（分钟）", "预计损失（元）"]
    for idx, (_, row) in enumerate(loss_df.iterrows(), start=2):
        apply_row_styles(ws, idx, detail_styles)
        for col, header in enumerate(loss_headers, 1):
            value = row.get(header)
            if header == "日期":
                value = parse_date(value)
                ws.cell(idx, col).number_format = "yyyy-mm-dd"
            elif header in {"异常时间（分钟）", "营业时长（分钟）", "预计损失（元）"}:
                value = round(to_number(value))
                ws.cell(idx, col).number_format = "#,##0"
            ws.cell(idx, col).value = value
    total_row_idx = 2 + len(loss_df)
    apply_row_styles(ws, total_row_idx, total_style)
    ws.cell(total_row_idx, 1).value = "合计"
    ws.cell(total_row_idx, 4).value = round(loss_df["异常时间（分钟）"].map(to_number).sum())
    ws.cell(total_row_idx, 5).value = round(loss_df["营业时长（分钟）"].map(to_number).sum())
    ws.cell(total_row_idx, 6).value = round(loss_df["预计损失（元）"].map(to_number).sum())
    for col in [4, 5, 6]:
        ws.cell(total_row_idx, col).number_format = "#,##0"
    summary_start = total_row_idx + 3
    for offset in range(7):
        apply_row_styles(ws, summary_start + offset, summary_styles[offset])
    copy_range_values(weekly_blocks, ws, 66, 2, 7, 6, summary_start, 1)
    ws.merge_cells(start_row=summary_start + 1, start_column=1, end_row=summary_start + 6, end_column=1)
    ws.merge_cells(start_row=summary_start + 1, start_column=2, end_row=summary_start + 3, end_column=2)
    ws.merge_cells(start_row=summary_start + 4, start_column=2, end_row=summary_start + 6, end_column=2)

    # Sheet6: CPC.
    ws = target_wb["CPC"]
    for row in [12, 36]:
        for start in [4, 7, 10, 13, 16]:
            ws.cell(row, start).value = current_dot
            ws.cell(row, start + 1).value = previous_dot
            ws.cell(row, start + 2).value = "周环比"
    clear_values(ws, 13, 17, 4, 18)
    clear_values(ws, 18, 32, 4, 18)
    clear_values(ws, 37, 41, 4, 18)
    clear_values(ws, 42, 56, 4, 18)
    ws.cell(1, 2).value = None
    clear_values(ws, 62, 62, 1, 18)

    mt_promo_cpc = mt_promo[
        mt_promo["_id"].isin(mt_ids)
        & ~mt_promo["场景"].astype(str).isin(MT_PROMO_EXCLUDE)
        & in_range(mt_promo, "_date", prev_start, current_end)
    ].copy()
    mt_cur, mt_prev = period_agg_rows(mt_promo_cpc, "_date", current_start, current_end, prev_start, prev_end)

    def mt_metrics(df):
        return {
            "spend": df["推广费"].map(to_number).sum(),
            "exposure": df["曝光次数"].map(to_number).sum(),
            "orders": df["总订单"].map(to_number).sum(),
            "revenue": df["交易额"].map(to_number).sum(),
        }

    def with_roi(cur_metrics, prev_metrics):
        return {
            "spend_cur": cur_metrics["spend"],
            "spend_prev": prev_metrics["spend"],
            "exposure_cur": cur_metrics["exposure"],
            "exposure_prev": prev_metrics["exposure"],
            "orders_cur": cur_metrics["orders"],
            "orders_prev": prev_metrics["orders"],
            "revenue_cur": cur_metrics["revenue"],
            "revenue_prev": prev_metrics["revenue"],
            "roi_cur": safe_ratio(cur_metrics["revenue"], cur_metrics["spend"]),
            "roi_prev": safe_ratio(prev_metrics["revenue"], prev_metrics["spend"]),
        }

    mt_group_rows = {"新客": 13, "菜品加速": 14, "老客": 15, "稳定转化": 16}
    for group, row_idx in mt_group_rows.items():
        cur_metrics = mt_metrics(mt_cur[mt_cur["场景"].astype(str) == group])
        prev_metrics = mt_metrics(mt_prev[mt_prev["场景"].astype(str) == group])
        write_metric_block(ws, row_idx, with_roi(cur_metrics, prev_metrics))
    mt_total_cur_metrics = mt_metrics(mt_cur)
    mt_total_prev_metrics = mt_metrics(mt_prev)
    write_metric_block(ws, 17, with_roi(mt_total_cur_metrics, mt_total_prev_metrics))
    for idx, store in enumerate(stores, start=18):
        ws.cell(idx, 1).value = int(store["mt_id"]) if store["mt_id"].isdigit() else store["mt_id"]
        ws.cell(idx, 2).value = store["name"]
        ws.cell(idx, 3).value = "合计"
        cur_metrics = mt_metrics(mt_cur[mt_cur["_id"] == store["mt_id"]])
        prev_metrics = mt_metrics(mt_prev[mt_prev["_id"] == store["mt_id"]])
        write_metric_block(ws, idx, with_roi(cur_metrics, prev_metrics))

    # Eleme CPC estimates promoted orders from visit lift, then applies store conversion and average income.
    ele_store_cpc = ele_store[ele_store["_id"].isin(ele_ids)].copy()
    store_ratio = {}
    for _, row in ele_store_cpc.iterrows():
        sid = id_text(row.get("门店编号"))
        d = row.get("_date")
        visit = to_number(row.get("进店人数"))
        orders = to_number(row.get("下单人数"))
        income = to_number(row.get("收入"))
        valid_orders = to_number(row.get("有效订单"))
        order_rate = orders / visit if visit else 0
        avg_income = income / valid_orders if valid_orders else 0
        store_ratio[(sid, d)] = (order_rate, avg_income)

    ele_promo_cpc = ele_promo[
        ele_promo["_id"].isin(ele_ids)
        & in_range(ele_promo, "_date", prev_start, current_end)
        & (ele_promo["推广产品"].astype(str) != "增量助手")
    ].copy()

    ele_promo_cpc["_group"] = ele_promo_cpc["推广产品"].astype(str).map(
        lambda v: "推广魔方" if "推广魔方" in v else ("一站式推广" if "一站式" in v else ("斗金推广" if "斗金" in v else v))
    )
    est_orders = []
    est_revenue = []
    for _, row in ele_promo_cpc.iterrows():
        sid = id_text(row.get("门店ID"))
        d = row.get("_date")
        order_rate, avg_income = store_ratio.get((sid, d), (0, 0))
        revenue_orders = to_number(row.get("进店提升数")) * ELE_VISIT_LIFT_TO_VISITOR_RATE * order_rate
        orders = revenue_orders
        revenue = revenue_orders * avg_income
        est_orders.append(orders)
        est_revenue.append(revenue)
    ele_promo_cpc["_est_orders"] = est_orders
    ele_promo_cpc["_est_revenue"] = est_revenue
    ele_cur, ele_prev = period_agg_rows(ele_promo_cpc, "_date", current_start, current_end, prev_start, prev_end)

    def ele_metrics(df):
        return {
            "spend": df["推广现金消费(元)"].map(to_number).sum(),
            "exposure": df["曝光提升数"].map(to_number).sum(),
            "orders": df["_est_orders"].map(to_number).sum(),
            "revenue": df["_est_revenue"].map(to_number).sum(),
        }

    ele_group_rows = {"推广魔方": 37, "增量助手": 38, "一站式推广": 39, "斗金推广": 40}
    for group, row_idx in ele_group_rows.items():
        cur_metrics = ele_metrics(ele_cur[ele_cur["_group"] == group])
        prev_metrics = ele_metrics(ele_prev[ele_prev["_group"] == group])
        write_metric_block(ws, row_idx, with_roi(cur_metrics, prev_metrics))
    ele_total_cur_metrics = ele_metrics(ele_cur)
    ele_total_prev_metrics = ele_metrics(ele_prev)
    write_metric_block(ws, 41, with_roi(ele_total_cur_metrics, ele_total_prev_metrics))
    for idx, store in enumerate(stores, start=42):
        ws.cell(idx, 1).value = int(store["ele_id"]) if store["ele_id"].isdigit() else store["ele_id"]
        ws.cell(idx, 2).value = store["name"]
        ws.cell(idx, 3).value = "合计"
        cur_metrics = ele_metrics(ele_cur[ele_cur["_id"] == store["ele_id"]])
        prev_metrics = ele_metrics(ele_prev[ele_prev["_id"] == store["ele_id"]])
        write_metric_block(ws, idx, with_roi(cur_metrics, prev_metrics))

    # Refresh the narrative panel on Sheet1 so the template does not retain old commentary.
    ws_overall = target_wb["整体业绩情况"]
    total_promo_spend = mt_total_cur_metrics["spend"] + ele_total_cur_metrics["spend"]
    total_promo_revenue = mt_total_cur_metrics["revenue"] + ele_total_cur_metrics["revenue"]
    total_promo_prev_revenue = mt_total_prev_metrics["revenue"] + ele_total_prev_metrics["revenue"]
    total_promo_roi = safe_ratio(total_promo_revenue, total_promo_spend)
    promo_revenue_delta = total_promo_revenue - total_promo_prev_revenue
    top_up = max(leaderboard_rows, key=lambda r: r[3] if r[3] is not None else -999)
    top_down = min(leaderboard_rows, key=lambda r: r[3] if r[3] is not None else 999)
    narrative = (
        "整体：\n"
        f"1、本周双平台营业额{fmt_money(weekly_summary.cell(3, 5).value)}，环比{fmt_pct(weekly_summary.cell(3, 7).value)}；"
        f"净收入{fmt_money(weekly_summary.cell(4, 5).value)}，环比{fmt_pct(weekly_summary.cell(4, 7).value)}；"
        f"有效订单{fmt_int(weekly_summary.cell(6, 5).value)}单，环比{fmt_pct(weekly_summary.cell(6, 7).value)}。\n"
        f"2、本周推广共计消耗{fmt_money(total_promo_spend)}，整体ROI为{fmt_roi(total_promo_roi)}，"
        f"推广带来实收{fmt_money(total_promo_revenue)}，环比{'增加' if promo_revenue_delta >= 0 else '减少'}{fmt_money(abs(promo_revenue_delta))}。\n\n"
        "门店分析：\n"
        f"①{top_up[0]}：本周实收{fmt_money(top_up[2])}，环比{fmt_pct(top_up[3])}，为本周增幅最高门店。\n"
        f"②{top_down[0]}：本周实收{fmt_money(top_down[2])}，环比{fmt_pct(top_down[3])}，为本周降幅最高门店。"
    )
    ws_overall["M4"].value = narrative
    ws_overall["M4"].alignment = Alignment(vertical="top", wrap_text=True)
    ws_overall.column_dimensions["C"].width = max(ws_overall.column_dimensions["C"].width or 0, 22)
    ws_overall.column_dimensions["M"].width = max(ws_overall.column_dimensions["M"].width or 0, 36)
    ws_overall.column_dimensions["N"].width = max(ws_overall.column_dimensions["N"].width or 0, 36)

    # Sheet7: trend append/update.
    ws = target_wb["业绩趋势"]
    trend_label = current_dot
    existing_col = None
    for col in range(2, ws.max_column + 1):
        if str(ws.cell(1, col).value).strip() == trend_label:
            existing_col = col
            break
    if existing_col is None:
        existing_col = ws.max_column + 1
        prev_col = existing_col - 1
        for row in [1, 2, 3]:
            copy_cell(ws.cell(row, prev_col), ws.cell(row, existing_col), copy_value=False, copy_style=True)
    ws.cell(1, existing_col).value = trend_label
    ws.cell(2, existing_col).value = to_number(weekly_summary.cell(4, 5).value)
    ws.cell(3, existing_col).value = to_number(weekly_summary.cell(6, 5).value)
    for row_idx in [1, 2, 3]:
        ws.cell(row_idx, existing_col).font = copy_font_with_overrides(ws.cell(row_idx, existing_col).font, name="微软雅黑")
    ws.cell(2, existing_col).number_format = "#,##0"
    ws.cell(3, existing_col).number_format = "#,##0"

    # A few readability safeguards.
    for sheet_name in ["中差评评价情况", "异常闭店情况"]:
        sh = target_wb[sheet_name]
        for row in sh.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = copy(cell.alignment)
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical="center",
                        wrap_text=cell.alignment.wrap_text,
                    )
    target_wb.calculation.fullCalcOnLoad = True
    target_wb.calculation.forceFullCalc = True
    clear_error_literals(target_wb)
    apply_qoq_style(target_wb)
    if callable(POSTPROCESS_HOOK):
        POSTPROCESS_HOOK(target_wb)
    target_wb.save(OUTPUT)

    print(f"saved={OUTPUT}")
    print(f"stores={len(stores)} current={current_full} previous={previous_full}")
    print(f"profit_target={profit_target:.2f} cumulative_profit={cumulative_profit:.2f}")
    print(f"products={len(rows)} reviews={len(review_df)} closure_rows={len(loss_df)}")


if __name__ == "__main__":
    main()
